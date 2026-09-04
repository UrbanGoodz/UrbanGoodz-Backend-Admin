import openpyxl
path = r"C:\Users\D'Andre Good\Downloads\urban_goodz_home_based_businesses_products_import_23_markup_FIXED_recommended.xlsx"
wb = openpyxl.load_workbook(path)
ws = wb.active
from collections import Counter
store_ids = [c.value for c in ws['R'][1:]]
cat_ids = [c.value for c in ws['E'][1:]]
print("total data rows:", ws.max_row-1)
print("StoreId counts:", Counter(store_ids))
print("CategoryId counts:", Counter(cat_ids))
print("distinct names sample per store:")
seen={}
for r in ws.iter_rows(min_row=2, values_only=True):
    seen.setdefault(r[17], []).append(r[1])
for s,names in seen.items():
    print(s, "->", len(names), "items:", names[:3])
