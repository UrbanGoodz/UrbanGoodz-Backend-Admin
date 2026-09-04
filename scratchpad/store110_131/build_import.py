import openpyxl
from openpyxl.styles import Font

# Template columns:
# Id, Name, Description, Image, CategoryId, SubCategoryId, UnitId, Stock, Price,
# Discount, DiscountType, AvailableTimeStarts, AvailableTimeEnds, Variations,
# ChoiceOptions, AddOns, Attributes, StoreId, ModuleId, Status, Veg, Recommended

COLUMNS = ["Id","Name","Description","Image","CategoryId","SubCategoryId","UnitId",
           "Stock","Price","Discount","DiscountType","AvailableTimeStarts",
           "AvailableTimeEnds","Variations","ChoiceOptions","AddOns","Attributes",
           "StoreId","ModuleId","Status","Veg","Recommended"]

MODULE = 6  # home-based

def prod(pid, name, desc, cat, stock, price, store):
    # SubCategoryId=0 (falsy -> controller falls back to CategoryId) to pass
    # the required-field check; UnitId blank; Image filename <=30 chars.
    return [pid, name, desc, "product.png", cat, 0, None, stock, price, 0,
            "amount", "08:00:00", "22:00:00", None, None, None, None, store,
            MODULE, "active", "no", "no"]

rows = []
pid = 3000

# =========================================================
# KEEP STORES (real vendor accounts, themed products)
# =========================================================

# 110 The Tipping Point (Latino) - Gifts & Snacks (70 Gifts / 71 Food)
keep110 = [
    ("Gourmet Snack Gift Box", "A curated gift box of snacks and pantry treats delivered locally.", 70, 18.99),
    ("Signature Hot Sauce", "House-made hot sauce with a slow-burn blend of peppers.", 71, 8.99),
    ("Artisan Coffee Sampler", "A sampler of locally roasted coffee beans for the perfect cup.", 71, 21.99),
    ("Celebration Gift Basket", "A ready-to-ship gift basket for birthdays and special occasions.", 70, 34.99),
    ("Handmade Greeting Cards (Set of 6)", "Blank greeting cards with local artisan art.", 70, 12.99),
    ("Pantry Essentials Bundle", "Everyday staples and seasonings in one convenient bundle.", 71, 27.99),
]
for n,d,c,p in keep110:
    rows.append(prod(pid,n,d,c,100,p,110)); pid+=1

# 111 Premium Goods Houston (Black, sneaker/streetwear) - Apparel 67
keep111 = [
    ("Premium Sneakers - Classic White", "Clean white premium sneakers for everyday style.", 67, 119.99),
    ("Limited Streetwear Tee", "Premium cotton streetwear tee with minimal branding.", 67, 39.99),
    ("Signature Hoodie", "Heavyweight hoodie with embroidered logo.", 67, 64.99),
    ("Premium Crew Socks (3-Pack)", "Cushioned crew socks in premium cotton.", 67, 16.99),
    ("Cap - Adjustable", "Adjustable cap with embroidered front patch.", 67, 24.99),
]
for n,d,c,p in keep111:
    rows.append(prod(pid,n,d,c,100,p,111)); pid+=1

# 115 Melodrama Boutique - Apparel 67 + Jewelry 68
keep115 = [
    ("Signature Floral Dress", "Flowing floral dress with a flattering fit.", 67, 78.99),
    ("Chunky Statement Necklace", "Bold chunky necklace to elevate any outfit.", 68, 34.99),
    ("Crop Top Set", "Matching crop top and skirt set in seasonal print.", 67, 52.99),
    ("Hoop Earrings - Gold Tone", "Everyday gold-tone hoop earrings.", 68, 14.99),
    ("Belted Trench Jacket", "Classic belted trench for every season.", 67, 89.99),
]
for n,d,c,p in keep115:
    rows.append(prod(pid,n,d,c,100,p,115)); pid+=1

# 117 Centre Dallas (Black) - Apparel 67 + Jewelry 68
keep117 = [
    ("Masterpiece Graphic Tee", "Graphic tee celebrating local art and culture.", 67, 32.99),
    ("Statement Blazer", "Modern statement blazer for bold styling.", 67, 94.99),
    ("Layered Necklace Set", "Delicate layered necklace set.", 68, 28.99),
    ("Wide-Leg Trousers", "Tailored wide-leg trousers in versatile neutral.", 67, 58.99),
    ("Tote Bag - Canvas", "Durable canvas tote with custom print.", 67, 22.99),
]
for n,d,c,p in keep117:
    rows.append(prod(pid,n,d,c,100,p,117)); pid+=1

# 118 Sneaker Politics (Black, sneakers) - Apparel 67
keep118 = [
    ("Signature Retro Sneaker", "Retro-inspired sneaker with premium cushioning.", 67, 109.99),
    ("High-Top Sneaker - Classic", "Classic high-top silhouette in leather.", 67, 99.99),
    ("Running Sneaker - Performance", "Lightweight running sneaker for daily miles.", 67, 89.99),
    ("Sneaker Socks (6-Pack)", "No-show and crew sneaker socks.", 67, 18.99),
    ("Shoe Care Kit", "Complete kit to keep your sneakers fresh.", 67, 26.99),
]
for n,d,c,p in keep118:
    rows.append(prod(pid,n,d,c,100,p,118)); pid+=1

# 126 League of Rebels (Black) - Apparel 67 + Print-On-Demand 83
keep126 = [
    ("Rebel Graphic Tee", "Graphic tee with bold slogan print.", 67, 34.99),
    ("Rebel Zip Hoodie", "Zip-up hoodie with chest print.", 67, 69.99),
    ("Custom Print Poster", "Custom print-on-demand poster of original art.", 83, 24.99),
    ("Rebel Snapback", "Snapback cap with raised embroidery.", 67, 27.99),
    ("Print-On-Demand T-Shirt", "Custom design printed on demand.", 83, 28.99),
]
for n,d,c,p in keep126:
    rows.append(prod(pid,n,d,c,100,p,126)); pid+=1

# =========================================================
# REPLACEMENT STORES (verified minority-owned, REAL products)
# =========================================================

# 112 LAMIK Beauty (Houston, Beauty 66)
lamik = [
    ("Custom Blend Foundation", "Custom-blended foundation for your exact skin tone.", 66, 49.00),
    ("Revelation Brow Duo", "Brow duo to define and shape.", 66, 49.00),
    ("Balancing Moisturizer", "Daily balancing moisturizer for all skin types.", 66, 29.00),
    ("Signature Lip Gloss", "High-shine lip gloss with nourishing oils.", 66, 18.00),
    ("Makeup Setting Spray", "Long-lasting setting spray.", 66, 24.00),
]
for n,d,c,p in lamik:
    rows.append(prod(pid,n,d,c,100,p,112)); pid+=1

# 113 Black Phlox Studios (Houston, Candles 76)
blackphlox = [
    ("Dapper Gentleman Candle", "Hand-poured soy candle with warm masculine notes.", 76, 24.29),
    ("Hustle Town Candle", "Hand-poured candle evoking hometown hustle.", 76, 24.29),
    ("Jewel Candle", "Hand-poured signature fragrance candle.", 76, 24.29),
    ("Room & Linen Spray", "Fragrance mist for rooms and linens.", 76, 18.00),
    ("Candle Gift Set", "Curated set of signature candles.", 76, 68.00),
]
for n,d,c,p in blackphlox:
    rows.append(prod(pid,n,d,c,100,p,113)); pid+=1

# 114 A Leap of Style (Houston, Apparel 67)
altstyle = [
    ("Africa is HOME Embroidered Sweatshirt", "Embroidered sweatshirt celebrating heritage.", 67, 45.00),
    ("Africa is HOME T-Shirt", "Comfort t-shirt with heritage print.", 67, 28.00),
    ("Black and Gold Silk Jacket", "Silk jacket in black and gold.", 67, 48.00),
    ("Heritage Tote Bag", "Tote bag with heritage print.", 67, 22.00),
    ("Statement Earrings", "Bold statement earrings.", 68, 19.00),
]
for n,d,c,p in altstyle:
    rows.append(prod(pid,n,d,c,100,p,114)); pid+=1

# 116 Custom Rings & Custom Things (Houston, Jewelry 68)
crct = [
    ("Omega Original XL Ring", "Handcrafted Omega-style XL ring.", 68, 525.00),
    ("Alpha Diagonal APhiA Ring", "Signature Alpha diagonal ring.", 68, 450.00),
    ("Delta Pyramid Bracelet", "Handcrafted pyramid-link bracelet.", 68, 75.00),
    ("Custom Name Ring", "Custom name ring made to order.", 68, 120.00),
    ("Pyramid Cuff Bracelet", "Handcrafted pyramid cuff.", 68, 95.00),
]
for n,d,c,p in crct:
    rows.append(prod(pid,n,d,c,100,p,116)); pid+=1

# 119 Charlene's Style Boutique (Dallas, Apparel 67)
charlene = [
    ("Booties Of Style - Black", "Stylish black booties for every outfit.", 67, 39.99),
    ("Chic On Chic Wedge Booties - Pink", "Chic pink wedge booties.", 67, 39.99),
    ("Suede Peep Open Toe Booties - Camel", "Suede camel peep-toe booties.", 67, 64.99),
    ("Trendy Top - Floral", "Trendy floral top.", 67, 24.99),
    ("Classic Handbag", "Classic structured handbag.", 67, 49.99),
]
for n,d,c,p in charlene:
    rows.append(prod(pid,n,d,c,100,p,119)); pid+=1

# 120 The House of Dasha (Dallas, Apparel 67)
dasha = [
    ("Jogger Sheer Dress - Hunter Green", "Sheer jogger dress in hunter green.", 67, 99.00),
    ("Roslyn Dress - Magenta", "Roslyn dress in bold magenta.", 67, 89.00),
    ("Off the Shoulder Peplum Dress", "Off-shoulder peplum dress.", 67, 45.00),
    ("Signature Top - Satin", "Satin signature top.", 67, 39.00),
    ("Statement Jumpsuit", "Tailored statement jumpsuit.", 67, 79.00),
]
for n,d,c,p in dasha:
    rows.append(prod(pid,n,d,c,100,p,120)); pid+=1

# 121 10 Hours of Fashion (Dallas, Apparel 67)
tenhours = [
    ("Camo Denim & Diamonds Blazer", "Camo denim blazer with diamond accents.", 67, 79.99),
    ("Leopard Street Chic Denim Vest", "Leopard print street-chic denim vest.", 67, 69.99),
    ("Sleeveless Statement Top", "Sleeveless top with statement print.", 67, 34.99),
    ("Curve-Fit Jumpsuit", "Curve-fit jumpsuit in seasonal print.", 67, 64.99),
    ("Custom Design Tee", "Custom-designed tee.", 67, 39.99),
]
for n,d,c,p in tenhours:
    rows.append(prod(pid,n,d,c,100,p,121)); pid+=1

# 122 Luminosa Vida (Austin, Candles 76 / Home Decor 72)
luminosa = [
    ("Divine Energy Candle", "Luxury coconut-wax candle for divine energy.", 76, 36.00),
    ("Night in Havana Candle", "Coconut-wax candle with leather, patchouli and tobacco notes.", 76, 28.80),
    ("Spiced Pear Candle", "Fall-inspired spiced pear candle.", 76, 36.00),
    ("Morning Light Reed Diffuser", "Reed diffuser for gentle fragrance.", 72, 40.00),
    ("Sacred Awakening Room Spray", "Room and linen spray for sacred space.", 72, 22.00),
]
for n,d,c,p in luminosa:
    rows.append(prod(pid,n,d,c,100,p,122)); pid+=1

# 123 Black Pearl Books (Austin, Books & Media 65)
bpb = [
    ("Parable of the Sower - Octavia E. Butler", "Classic sci-fi novel by Octavia E. Butler.", 65, 19.99),
    ("All About Love - bell hooks", "bell hooks on love and community.", 65, 17.99),
    ("Transcendent Kingdom - Yaa Gyasi", "Critically acclaimed novel by Yaa Gyasi.", 65, 17.00),
    ("Black Literature Bestseller", "Featured bestselling Black literature title.", 65, 18.99),
    ("Bookmark Gift Set", "Set of art bookmarks.", 65, 12.99),
]
for n,d,c,p in bpb:
    rows.append(prod(pid,n,d,c,100,p,123)); pid+=1

# 124 U4U Designs (Austin, Apparel 67)
u4u = [
    ("Emem Dress", "Signature designed dress.", 67, 75.00),
    ("Yellow and Black Tire Dye Wrap Skirt", "Tire-dye wrap skirt in yellow and black.", 67, 89.00),
    ("Brown Bomber Jacket", "Premium brown bomber jacket.", 67, 180.00),
    ("Signature Ankara Dress", "Ankara print signature dress.", 67, 95.00),
    ("Matching Headwrap", "Matching Ankara headwrap.", 67, 25.00),
]
for n,d,c,p in u4u:
    rows.append(prod(pid,n,d,c,100,p,124)); pid+=1

# 125 Floral Sea (Austin, Jewelry 68)
floralsea = [
    ("Signature Ball Cuff Bracelet - Rainbow", "Ball cuff bracelet in rainbow accents.", 68, 69.00),
    ("Signature Beaded Choker - Rainbow Blooms", "Beaded choker in rainbow blooms.", 68, 127.00),
    ("Criss-Cross Ring - Rainbow", "Criss-cross ring in rainbow.", 68, 45.00),
    ("Layered Necklace", "Layered statement necklace.", 68, 55.00),
]
for n,d,c,p in floralsea:
    rows.append(prod(pid,n,d,c,100,p,125)); pid+=1

# 127 Lady Brown's Boutique (Galveston, Apparel 67 + Jewelry 68)
ladybrown = [
    ("Dana Dress", "Signature Dana dress.", 67, 46.99),
    ("Rose Bottoms", "Rose print bottoms.", 67, 46.99),
    ("Studded Fanny Pack", "Studded fanny pack.", 68, 18.99),
    ("Custom Tee - Graphic", "Custom graphic tee.", 67, 24.99),
    ("Statement Necklace", "Statement necklace.", 68, 22.99),
]
for n,d,c,p in ladybrown:
    rows.append(prod(pid,n,d,c,100,p,127)); pid+=1

# 128 Cloth & Cord (Kemah, Austin none; Galveston County, Jewelry 68)
clothcord = [
    ("African Cloth Earrings - Red Ankara Hoop", "Red Ankara African cloth hoop earrings.", 68, 22.00),
    ("African Necklace - Red Ankara", "Red Ankara African necklace.", 68, 35.00),
    ("African Choker - Kente/Ankara", "Kente and Ankara African choker.", 68, 35.00),
    ("Ankara Bracelet Set", "Set of Ankara cloth bracelets.", 68, 28.00),
    ("African Print Statement Necklace", "Bold African print necklace.", 68, 38.00),
]
for n,d,c,p in clothcord:
    rows.append(prod(pid,n,d,c,100,p,128)); pid+=1

# 129 BLCK Market (Pearland, Brazoria; marketplace - Apparel 67 + Handmade 69)
blckmarket = [
    ("Shop Local Tee", "Tee celebrating shopping local.", 67, 28.00),
    ("Boyfriend Jeans", "Relaxed boyfriend jeans.", 67, 34.99),
    ("3 Piece Cammy Set - Oatmeal", "Three-piece cammy set in oatmeal.", 67, 49.99),
    ("Handmade Tote Bag", "Handmade canvas tote.", 69, 24.00),
    ("BLCK Market Cap", "Adjustable cap with BLCK Market logo.", 67, 22.00),
]
for n,d,c,p in blckmarket:
    rows.append(prod(pid,n,d,c,100,p,129)); pid+=1

# 130 Ebony Expressions (Huntsville, Art 73 / Stationery 77 / Print 83)
ebony = [
    ("Be Still Wall Art", "Printable wall art - Be Still and Know.", 73, 9.99),
    ("For I Know the Plans Wall Art", "Printable wall art - encouraging verse.", 73, 9.99),
    ("I Can Do All Things Wall Art", "Printable wall art - affirmation.", 73, 9.99),
    ("Faith Bookmark Set", "Set of faith-themed bookmarks.", 77, 7.99),
    ("Inspirational Print-On-Demand Poster", "Print-on-demand inspirational poster.", 83, 14.99),
]
for n,d,c,p in ebony:
    rows.append(prod(pid,n,d,c,100,p,130)); pid+=1

# 131 (Huntsville, 2nd) NOheartNOhustle (Apparel 67 + Print 83)
noheart = [
    ("NOheartNOhustle Signature Tee", "Signature graphic tee.", 67, 29.99),
    ("NOheartNOhustle Hoodie", "Heavyweight hoodie with signature print.", 67, 59.99),
    ("Hustle Print-On-Demand Tee", "Print-on-demand tee with hustle slogan.", 83, 26.99),
    ("NOheartNOhustle Cap", "Snapback with embroidery.", 67, 24.99),
    ("Motivational Print Poster", "Print-on-demand motivational poster.", 83, 18.99),
]
for n,d,c,p in noheart:
    rows.append(prod(pid,n,d,c,100,p,131)); pid+=1

# =========================================================
# Write Excel
# =========================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"
ws.append(COLUMNS)
for c in COLUMNS:
    ws.cell(row=1, column=COLUMNS.index(c)+1).font = Font(bold=True)
for r in rows:
    ws.append(r)

from collections import Counter

out = r"C:\Users\D'Andre Good\Documents\GitHub\AdminPanel_Update_V39\scratchpad\store110_131\urban_goodz_home_based_businesses_110_131_import.xlsx"
wb.save(out)
print("Saved full file:", out)

# ---- Validation: mimic the controller's required-field checks ----
def validate(r):
    checks = [r[0], r[1], r[4], r[5], r[8], r[17], r[18], r[9], r[10]]
    empty = [c for c in checks if c == '' or c is None]
    if empty:
        return f"FAIL empty required Id={r[0]}"
    if r[8] < 0:
        return f"FAIL price<0 Id={r[0]}"
    if r[9] < 0:
        return f"FAIL discount<0 Id={r[0]}"
    if r[3] and len(str(r[3])) > 30:
        return f"FAIL image>30chars Id={r[0]}"
    if r[11] > r[12]:
        return f"FAIL times Id={r[0]}"
    return None

bad = [validate(r) for r in rows]
bad = [b for b in bad if b]
print("Validation errors:", bad if bad else "NONE - all rows pass controller checks")

# ---- Test batch: only stores 111 and 118 (2 stores) ----
test_stores = {111, 118}
test_rows = [r for r in rows if r[17] in test_stores]
tb = openpyxl.Workbook()
ts = tb.active
ts.title = "Sheet1"
ts.append(COLUMNS)
for c in COLUMNS:
    ts.cell(row=1, column=COLUMNS.index(c)+1).font = Font(bold=True)
for r in test_rows:
    ts.append(r)
test_out = r"C:\Users\D'Andre Good\Documents\GitHub\AdminPanel_Update_V39\scratchpad\store110_131\urban_goodz_TEST_batch_stores_111_118.xlsx"
tb.save(test_out)
print("Saved test file:", test_out, "rows:", len(test_rows), "stores:", sorted(test_stores))
print("Test file StoreIds:", dict(Counter(r[17] for r in test_rows)))
